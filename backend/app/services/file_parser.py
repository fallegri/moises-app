"""File parser service for extracting text from uploaded documents."""

import io
from pathlib import Path
from typing import Union

from docx import Document as DocxDocument
from openpyxl import load_workbook


class FileParserService:
    """Service for parsing uploaded files and extracting text content."""

    SUPPORTED_EXTENSIONS = {".docx", ".xlsx", ".md", ".txt"}

    def parse_file(self, file_content: bytes, filename: str) -> str:
        """Parse a file and extract its text content.

        Args:
            file_content: Raw file bytes.
            filename: Original filename to determine format.

        Returns:
            Extracted text content.

        Raises:
            ValueError: If file format is not supported.
        """
        extension = Path(filename).suffix.lower()

        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file format: {extension}. "
                f"Supported formats: {', '.join(self.SUPPORTED_EXTENSIONS)}"
            )

        if extension == ".docx":
            return self._parse_docx(file_content)
        elif extension == ".xlsx":
            return self._parse_xlsx(file_content)
        elif extension in (".md", ".txt"):
            return self._parse_text(file_content)

        raise ValueError(f"Unsupported file format: {extension}")

    def _parse_docx(self, content: bytes) -> str:
        """Extract text from a .docx file."""
        doc = DocxDocument(io.BytesIO(content))
        paragraphs = []
        for paragraph in doc.paragraphs:
            text = paragraph.text.strip()
            if text:
                paragraphs.append(text)
        return "\n\n".join(paragraphs)

    def _parse_xlsx(self, content: bytes) -> str:
        """Extract text from a .xlsx file."""
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        text_parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Hoja: {sheet_name} ---")

            for row in ws.iter_rows(values_only=True):
                row_values = []
                for cell in row:
                    if cell is not None:
                        row_values.append(str(cell))
                if row_values:
                    text_parts.append(" | ".join(row_values))

        wb.close()
        return "\n".join(text_parts)

    def _parse_text(self, content: bytes) -> str:
        """Extract text from a plain text or markdown file."""
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            return content.decode("latin-1")
